# -*- coding: utf-8 -*-
import os
import re
import openpyxl
import importlib
from google.protobuf import text_format
from common.genpb import GenAsPython

CPP_TYPE = {}
CPP_TYPE[1] = int
CPP_TYPE[2] = int
CPP_TYPE[3] = int
CPP_TYPE[4] = int
CPP_TYPE[5] = float
CPP_TYPE[6] = float
CPP_TYPE[7] = bool
# 8: enumerate
CPP_TYPE[9] = str
# 10: message

PROTO_TYPE = {}
PROTO_TYPE[1] = 'optional'
PROTO_TYPE[2] = 'required'
PROTO_TYPE[3] = 'repeated'

CONV_MODULE = {}

__rootPath = os.path.dirname(os.path.abspath(__file__))
__pypbPath = __rootPath + '/./pypb/'
__convRe = r'Conv\(([\w\d_]+)\.proto\, *([\w\d_]+)\)'
__convPy = 'common.pypb.{}_pb2'


def __loadOne(convKeys, field, val):
    print(field.default_value, val, field.cpp_type)
    val = val or field.default_value
    if field.cpp_type in CPP_TYPE:
        if (field.cpp_type == 1 or field.cpp_type == 2) and type(val) == type(1.0):
            val = CPP_TYPE[field.cpp_type](round(val))
        elif (field.cpp_type == 1 or field.cpp_type == 2 or field.cpp_type == 3 or field.cpp_type == 4):
            try:
                val = CPP_TYPE[field.cpp_type](round(float(val)))
            except:
                tmpVal = val.decode('utf-8')
                if tmpVal not in convKeys:
                    print(tmpVal + ' not in convKeys')
                    quit()
                val = convKeys[tmpVal]
        elif field.cpp_type == 9 and type(val) == type(b''):  # str
            val = val.decode('utf-8')
        else:
            val = CPP_TYPE[field.cpp_type](val)
    return val


def __loadField(convKeys, field, convRow, val):
    # repeated
    # print(field.name, field.label)
    key = field.name
    if field.label == 3:
        oneObj = __loadOne(convKeys, field, val)
        # print(oneObj)
        getattr(convRow, key).append(oneObj)
    else:
        # not message
        if field.cpp_type != 10:
            oneObj = __loadOne(convKeys, field, val)
            convRow.__setattr__(key, oneObj)
        else:
            oneObj = __loadOne(convKeys, field, val)
            fieldIns = convRow.__getattribute__(key)
            text_format.Merge(oneObj, fieldIns)
    # print()


def __splitKey(protoType, key):
    if key not in protoType.fields_by_name:
        ds = key.split('.')
        if len(ds) == 1 or ds[0] not in protoType.fields_by_name:
            return None, None
        return ds[0], '.'.join(ds[1:len(ds)])
    return key, None


def __innerArray(field, subKey):
    if '.' in subKey:
        subKey, subSubKey = __splitKey(field.message_type, subKey)
        subField = field.message_type.fields_by_name[subKey]
        if subField.label == 3:
            return True
        return __innerArray(subField, subSubKey)
    else:
        subField = field.message_type.fields_by_name[subKey]
        if subField.label == 3:
            return True
    return False


def __loadMsg(convKeys, convModule, fieldModule, convRow, key, val, repeated, prefix=''):
    protoType = fieldModule.message_type
    key, subKey = __splitKey(protoType, key)
    if key == None:
        return None
    field = protoType.fields_by_name[key]
    if field.message_type == None or subKey == None:
        return __loadField(convKeys, field, convRow, val)
    else:
        fullKey = prefix + '.' + key
        if field.label == 3:
            if __innerArray(field, subKey):
                subConvRow = getattr(convRow, key)[-1]
            else:
                if fullKey in repeated:
                    subConvRow = repeated[fullKey]
                else:
                    subConvRow = getattr(convRow, key).add()
                    repeated[fullKey] = subConvRow
        else:
            subConvRow = getattr(convRow, key)
        return __loadMsg(convKeys, convModule, field, subConvRow, subKey, val, repeated, fullKey)


def __loadConvModule(convFile):
    convModule = __convPy.format(convFile)
    return importlib.import_module(convModule)


def __writeFile(exportPath, data, name, suffix, mode='wb+'):
    with open(exportPath + name + '.' + suffix, mode) as file:
        file.write(data)
        file.close()


def __convXlsx(convKeys, exportPath, filepath, xlsxname):
    print('convert file:' + filepath)
    wb = openpyxl.load_workbook(filepath)
    # print(wb.sheetnames)
    for sheetname in wb.sheetnames:
        print('export:{}; sheet:{}'.format(xlsxname, sheetname))
        sheet = wb[sheetname]
        # 转换的函数
        convStr = sheet.cell(1, 1).value
        convSp = re.findall(__convRe, convStr)[0]
        convSpLen = len(convSp)
        if convSpLen != 2:
            print('sheet:{}; convstr:{}; len:{}; invaild'.format(
                sheetname, convStr, convSpLen))
            continue
        # 这个文件的pypb
        convFile, convMsg = convSp[0], convSp[1]
        convMsgRows = convMsg + 'Rows'
        convModule = __loadConvModule(convFile)
        convModuleRows = convModule.__dict__[convMsgRows]()
        # protoType = convModule.DESCRIPTOR.message_types_by_name[convMsg]
        protoTypeRows = convModule.DESCRIPTOR.message_types_by_name[convMsgRows]
        convOneMsg = protoTypeRows.fields[0]
        if sheet[4][0].value.lower() != 'id':
            print('first key not id')
            quit()
        for rowid in range(5, sheet.max_row + 1):
            repeated = {}
            for colid in range(sheet.max_column):
                key = sheet[4][colid].value
                val = sheet[rowid][colid].value
                if val == None or val == '':
                    continue
                if type(val) == type(u''):
                    val = val.encode('utf8')
                # 首行不是空的时候新add
                if colid == 0 and val != None:
                    convRows = getattr(convModuleRows, convOneMsg.name)
                    convRow = convRows.add()
                __loadMsg(convKeys, convModule, convOneMsg,
                          convRow, key, val, repeated)

        __writeFile(exportPath, convModuleRows.SerializePartialToString(),
                    convMsg, 'pbin')
        __writeFile(exportPath, text_format.MessageToString(
            convModuleRows), convMsg, 'txt', 'w+')
        print('\n')


def GenPbin(convKeys, protocPath, scanBasePath, protoPath, excelPath, exportPath, includePath):
    GenAsPython(protocPath, scanBasePath, protoPath, __pypbPath, includePath)
    if not os.path.exists(excelPath):
        print(excelPath + ' not exist')
        return
    print(excelPath + ' convert start')
    list = os.listdir(excelPath)
    for line in list:
        filepath = os.path.join(excelPath, line)
        if os.path.isfile(filepath) and os.path.splitext(filepath)[-1] == '.xlsx':
            __convXlsx(convKeys, exportPath, filepath, line)
