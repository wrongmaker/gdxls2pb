# -*- coding: utf-8 -*-
import os
import re
import openpyxl
import importlib

__convRe = r'Conv\(([\w\d_]+)\.proto\, *([\w\d_]+)\)'


def __convTypeCpp(ktype):
    if ktype == 'num':
        return 'int32'
    elif ktype == 'bignum':
        return 'int64'
    elif ktype == 'str':
        return 'string'
    elif ktype == 'float':
        return 'float'
    elif ktype == 'ffloat':
        return 'double'
    return 'int32'


def __writeFile(protoPath, data, name, suffix, mode='w+'):
    with open(protoPath + name + '.' + suffix, mode) as file:
        file.write(data)
        file.close()


def __convXlsx(protoPath, filepath, xlsxname):
    print('generate file:' + filepath)
    outputData = ''
    wb = openpyxl.load_workbook(filepath)
    for sheetname in wb.sheetnames:
        print('generate:{}; sheet:{}'.format(xlsxname, sheetname))
        sheet = wb[sheetname]
        # 转换的函数
        convStr = sheet.cell(1, 1).value
        convSp = re.findall(__convRe, convStr)[0]
        protoFile = convSp[0]
        protoName = convSp[1]
        protoNameRows = protoName + 'Rows'
        print(protoFile, protoName)
        if len(outputData) == 0:
            outputData = 'syntax = "proto2";\n\npackage {};\n\n'.format(protoFile)

        tmpOutputData = "// {} \n".format(sheet.title)
        tmpOutputData += "message {} {{\n".format(protoName)
        repeatedCols = sheet[1][2].value or ''
        repeatedColsList = repeatedCols.split(';')
        optionalrows = 1
        otherStructMap = {}
        for colid in range(sheet.max_column):
            desc = sheet[2][colid].value
            ktype = sheet[3][colid].value
            key = sheet[4][colid].value
            if key == None:
                continue
            desc = desc.replace('\r\n', '').replace('\n', '') 
            typeCpp = __convTypeCpp(ktype)
            repeatedStruct = key.split(".")
            if len(repeatedStruct) > 1:
                pbmsgtype = 'optional'
                if key in repeatedColsList:
                    pbmsgtype = 'repeated'
                if repeatedStruct[0] in otherStructMap:
                    otherStructMap[repeatedStruct[0]].append(
                        [pbmsgtype, typeCpp, repeatedStruct[1], desc])
                    continue
                else:
                    otherStructMap[repeatedStruct[0]] = [
                        [pbmsgtype, typeCpp, repeatedStruct[1], desc]]
                    typeCpp = protoName + repeatedStruct[0] + 'Cfg'
                    key = repeatedStruct[0]
                    desc = (desc or '') + ',...'
            pbmsgtype = 'optional'
            if key in repeatedColsList:
                pbmsgtype = 'repeated'
            tmpOutputData += "  {} {} {} = {};  // {}\n".format(
                pbmsgtype, typeCpp, key, optionalrows, desc)
            optionalrows += 1

        tmpOutputData += "}\n\n"
        tmpOutputData += "message {} {{\n  repeated {} rows = 1;\n}}\n\n".format(
            protoNameRows, protoName)

        for key, val in otherStructMap.items():
            outputData += "message {} {{\n".format(protoName + key + 'Cfg')
            optionalrows = 1
            for vkey in val:
                outputData += "  {} {} {} = {};  // {}\n".format(
                    vkey[0], vkey[1], vkey[2], optionalrows, vkey[3])
                optionalrows += 1
            outputData += "}\n\n"
        outputData += tmpOutputData

    print(outputData)
    print('\n')
    __writeFile(protoPath, outputData, protoFile, 'proto')


def GenProto(excelPath, protoPath):
    if not os.path.exists(excelPath):
        return
    list = os.listdir(excelPath)
    for line in list:
        filepath = os.path.join(excelPath, line)
        if os.path.isfile(filepath) and os.path.splitext(filepath)[-1] == '.xlsx':
            __convXlsx(protoPath, filepath, line)
